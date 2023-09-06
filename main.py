import openpyxl as xl
import pandas as pd
from datetime import datetime
from io import StringIO
import requests

def main():

    # get the excel file from ndx_data
    excel_file_url = "https://stooq.com/q/d/l/?s=^ndx&i=d"

    response = requests.get(excel_file_url)

    response.raise_for_status()

    npx_data = pd.read_csv(StringIO(response.text))


    # open excel workbook
    main_wb = xl.load_workbook("resources\\new.xlsx")

    sheet = main_wb["Rolling Returns"]    


    # overrite old data
    amount_changed = 0
    num_trading_days = 0
    curr_index = 3
    for date, close in zip(npx_data["Date"], npx_data["Close"]):
        # date (K)
        sheet.cell(curr_index, 11, datetime.strptime(date, "%Y-%m-%d"))

        # amount (L)
        sheet.cell(curr_index, 12, float(close))

        # since we're using the next trading day as our lookback
        if num_trading_days >= 253:
            # start adding formula to column N
            price_lookback_formula = f"={sheet.cell(curr_index, 12).coordinate} / INDEX(L:L, MAX(ROW({sheet.cell(curr_index, 12).coordinate}) - 252, 3)) - 1"

            tr_lookback_formula = f"={sheet.cell(curr_index, 13).coordinate} / INDEX(L:L, MAX(ROW({sheet.cell(curr_index, 13).coordinate}) - 252, 3)) - 1"
            
            # Column N
            sheet.cell(curr_index, 14, price_lookback_formula)
            sheet.cell(curr_index, 15, tr_lookback_formula)

        num_trading_days += 1
        curr_index += 1
        amount_changed += 1

    # values changed
    print(f'Expected: {len(npx_data["Close"])}\nActual: {amount_changed}')
    
    # save workbook
    main_wb.save("resources\\new.xlsx")
    return 0



if __name__ == "__main__":
    main()