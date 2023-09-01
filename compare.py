import openpyxl as xl



if __name__ == "__main__":
    new_wb = xl.load_workbook("resources\\new.xlsx")
    old_wb = xl.load_workbook("resources\\old.xlsx")


    new_sheet = new_wb["Rolling Returns"]
    old_sheet = old_wb["Rolling Returns"]
    num = 3
    for row in new_sheet.iter_rows(min_row=3, max_row=new_sheet.max_row, min_col=12, max_col=12):
            if row[0].value == old_sheet.cell(num, 12).value:
                  print("Difference found")
            num += 1